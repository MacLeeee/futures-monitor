#!/usr/bin/env python3
# ============================================================
# 期货交易历史导出工具 - 生成 Excel 报表
# 使用方法：python3 scripts/export_positions_excel.py
# ============================================================

import json
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
POSITIONS_FILE = ROOT / "futures-monitor" / "public" / "positions.json"
OUTPUT_DIR = ROOT
OUTPUT_FILE = OUTPUT_DIR / f"交易历史_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ── 配色 ────────────────────────────────────────────────────
COLOR_HEADER    = "1F4E79"   # 深蓝 - 表头
COLOR_SUBHEADER = "2E75B6"   # 中蓝 - 副标题
COLOR_PROFIT    = "C6EFCE"   # 浅绿 - 盈利
COLOR_LOSS      = "FFCCCC"   # 浅红 - 亏损
COLOR_OPEN      = "FFF2CC"   # 浅黄 - 持仓中
COLOR_LONG      = "E2EFDA"   # 淡绿 - 做多行
COLOR_SHORT     = "FCE4D6"   # 淡橙 - 做空行
COLOR_ALT_ROW   = "F5F5F5"   # 浅灰 - 间隔行
COLOR_SUMMARY   = "D6E4F0"   # 汇总区背景

STATUS_MAP = {
    "open":      "持仓中",
    "closed_tp": "止盈出场",
    "closed_sl": "止损出场",
}
DIR_MAP = {
    "long":  "做多",
    "short": "做空",
}
SIG_MAP = {
    "breakout": "突破",
    "pullback": "回踩",
    "box":      "箱体",
}


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def cell_style(ws, row, col, value, *,
               bold=False, font_color="000000", fill=None,
               align="center", num_fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=font_color, name="微软雅黑", size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if fill:
        c.fill = fill
    if num_fmt:
        c.number_format = num_fmt
    if border:
        c.border = thin_border()
    return c


def write_sheet_all(wb: openpyxl.Workbook, positions: list[dict]):
    """全部交易记录Sheet"""
    ws = wb.active
    ws.title = "全部交易记录"
    ws.freeze_panes = "A3"   # 冻结前两行

    # ── 标题行 ──────────────────────────────────────────────
    ws.merge_cells("A1:R1")
    c = ws.cell(row=1, column=1, value="期货交易历史总览")
    c.font = Font(bold=True, size=14, color="FFFFFF", name="微软雅黑")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = make_fill(COLOR_HEADER)
    ws.row_dimensions[1].height = 30

    # ── 列定义 ──────────────────────────────────────────────
    headers = [
        ("序号",     5),  ("交易ID",    22), ("品种",    8),
        ("方向",     7),  ("信号类型",  9),  ("入场时间",15),
        ("入场价",   12), ("止损价",   12), ("止盈价",  12),
        ("风险距离", 10), ("ATR",      10), ("出场时间",15),
        ("出场价",   12), ("盈亏(点)", 12), ("盈亏(%)", 10),
        ("状态",      9), ("是否移动止损", 12), ("备注", 10),
    ]
    hdr_fill = make_fill(COLOR_SUBHEADER)
    for col_i, (name, width) in enumerate(headers, 1):
        cell_style(ws, 2, col_i, name, bold=True, font_color="FFFFFF",
                   fill=hdr_fill, align="center")
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[2].height = 22

    # ── 数据行 ──────────────────────────────────────────────
    sorted_pos = sorted(positions,
                        key=lambda p: p.get("entryTime", ""),
                        reverse=True)
    for i, pos in enumerate(sorted_pos, 1):
        row = i + 2
        status  = pos.get("status", "")
        direc   = pos.get("direction", "")
        pnl     = pos.get("pnl")
        pnl_pct = pos.get("pnlPct")

        # 底色逻辑：已平仓按盈亏色，持仓中用黄色，间隔用灰
        if status == "open":
            row_fill = make_fill(COLOR_OPEN)
        elif pnl is not None and pnl > 0:
            row_fill = make_fill(COLOR_PROFIT)
        elif pnl is not None and pnl < 0:
            row_fill = make_fill(COLOR_LOSS)
        else:
            row_fill = make_fill(COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF")

        def rc(col, val, num_fmt=None, align="center"):
            cell_style(ws, row, col, val, fill=row_fill, align=align, num_fmt=num_fmt)

        rc(1,  i)
        rc(2,  pos.get("id", ""),       align="left")
        rc(3,  pos.get("symbol", ""))
        rc(4,  DIR_MAP.get(direc, direc))
        rc(5,  SIG_MAP.get(pos.get("signalType", ""), pos.get("signalType", "")))
        rc(6,  pos.get("entryTime", ""))
        rc(7,  pos.get("entryPrice"),   num_fmt="#,##0.00##")
        rc(8,  pos.get("stopLoss"),     num_fmt="#,##0.00##")
        rc(9,  pos.get("takeProfit"),   num_fmt="#,##0.00##")
        rc(10, pos.get("riskDist"),     num_fmt="0.0000")
        rc(11, pos.get("atr"),          num_fmt="0.0000")
        rc(12, pos.get("exitTime", ""))
        rc(13, pos.get("exitPrice"),    num_fmt="#,##0.00##")
        rc(14, pnl,                     num_fmt="+#,##0.00##;-#,##0.00##")
        rc(15, (pnl_pct / 100) if pnl_pct is not None else None,
                                        num_fmt="+0.00%;-0.00%")
        rc(16, STATUS_MAP.get(status, status))
        rc(17, "是" if pos.get("trailingActive") else "否")
        rc(18, "")

        # 盈亏列特殊字色
        pnl_cell     = ws.cell(row=row, column=14)
        pnl_pct_cell = ws.cell(row=row, column=15)
        if pnl is not None:
            color = "006100" if pnl > 0 else ("9C0006" if pnl < 0 else "000000")
            pnl_cell.font     = Font(bold=True, color=color, name="微软雅黑", size=10)
            pnl_pct_cell.font = Font(bold=True, color=color, name="微软雅黑", size=10)

        ws.row_dimensions[row].height = 18


def write_sheet_summary(wb: openpyxl.Workbook, positions: list[dict]):
    """统计汇总Sheet"""
    ws = wb.create_sheet("统计汇总")

    closed = [p for p in positions if p["status"] != "open"]
    wins   = [p for p in closed if (p.get("pnl") or 0) > 0]
    losses = [p for p in closed if (p.get("pnl") or 0) < 0]
    opens  = [p for p in positions if p["status"] == "open"]

    total_pnl  = sum(p.get("pnl") or 0 for p in closed)
    win_pnl    = sum(p.get("pnl") or 0 for p in wins)
    loss_pnl   = sum(p.get("pnl") or 0 for p in losses)
    win_rate   = len(wins) / len(closed) if closed else 0

    rows_data = [
        ("总交易笔数",      len(positions),   None),
        ("已平仓笔数",      len(closed),      None),
        ("持仓中笔数",      len(opens),       None),
        ("盈利笔数",        len(wins),        None),
        ("亏损笔数",        len(losses),      None),
        ("胜率",            win_rate,         "0.00%"),
        ("总盈亏(点)",      total_pnl,        "+#,##0.00;-#,##0.00"),
        ("总盈利(点)",      win_pnl,          "+#,##0.00;-#,##0.00"),
        ("总亏损(点)",      loss_pnl,         "+#,##0.00;-#,##0.00"),
        ("平均盈利(点)",    win_pnl  / len(wins)   if wins   else 0, "#,##0.00"),
        ("平均亏损(点)",    loss_pnl / len(losses) if losses else 0, "#,##0.00"),
        ("盈亏比",          abs(win_pnl/len(wins)) / abs(loss_pnl/len(losses))
                            if wins and losses else 0,  "0.00"),
    ]

    ws.merge_cells("A1:C1")
    c = ws.cell(row=1, column=1, value="交易统计汇总")
    c.font      = Font(bold=True, size=13, color="FFFFFF", name="微软雅黑")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = make_fill(COLOR_HEADER)
    ws.row_dimensions[1].height = 28

    hdrs = ["指标", "数值", "备注"]
    hdr_fill = make_fill(COLOR_SUBHEADER)
    for ci, h in enumerate(hdrs, 1):
        cell_style(ws, 2, ci, h, bold=True, font_color="FFFFFF",
                   fill=hdr_fill, align="center")

    summary_fill = make_fill(COLOR_SUMMARY)
    for ri, (label, value, fmt) in enumerate(rows_data, 3):
        cell_style(ws, ri, 1, label, fill=summary_fill, align="left", bold=True)
        vc = ws.cell(row=ri, column=2, value=value)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border    = thin_border()
        vc.fill      = make_fill("FFFFFF")
        if fmt:
            vc.number_format = fmt
        # 盈亏行上色
        if "盈亏" in label or "盈利" in label or "亏损" in label:
            color = "006100" if (value or 0) > 0 else ("9C0006" if (value or 0) < 0 else "000000")
            vc.font = Font(bold=True, color=color, name="微软雅黑", size=10)
        cell_style(ws, ri, 3, "", fill=make_fill("FFFFFF"))
        ws.row_dimensions[ri].height = 20

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    # ── 按信号类型小计 ──────────────────────────────────────
    start_row = len(rows_data) + 5

    ws.merge_cells(f"A{start_row}:C{start_row}")
    c2 = ws.cell(row=start_row, column=1, value="按信号类型统计")
    c2.font      = Font(bold=True, size=11, color="FFFFFF", name="微软雅黑")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill      = make_fill(COLOR_SUBHEADER)
    ws.row_dimensions[start_row].height = 22

    sub_hdrs = ["信号类型", "笔数", "净盈亏(点)", "胜率"]
    for ci, h in enumerate(sub_hdrs, 1):
        cell_style(ws, start_row + 1, ci, h, bold=True, font_color="FFFFFF",
                   fill=make_fill(COLOR_SUBHEADER))

    sig_types = {"breakout": "突破", "pullback": "回踩", "box": "箱体"}
    for ri_off, (sig_key, sig_name) in enumerate(sig_types.items(), start_row + 2):
        grp     = [p for p in closed if p.get("signalType") == sig_key]
        grp_win = [p for p in grp if (p.get("pnl") or 0) > 0]
        grp_pnl = sum(p.get("pnl") or 0 for p in grp)
        grp_wr  = len(grp_win) / len(grp) if grp else 0
        alt = make_fill(COLOR_ALT_ROW if ri_off % 2 == 0 else "FFFFFF")
        cell_style(ws, ri_off, 1, sig_name, fill=alt, bold=True, align="center")
        c_cnt = ws.cell(row=ri_off, column=2, value=len(grp))
        c_cnt.alignment = Alignment(horizontal="right", vertical="center")
        c_cnt.fill = alt; c_cnt.border = thin_border()
        c_pnl = ws.cell(row=ri_off, column=3, value=grp_pnl)
        c_pnl.number_format = "+#,##0.00;-#,##0.00"
        c_pnl.alignment = Alignment(horizontal="right", vertical="center")
        c_pnl.fill = alt; c_pnl.border = thin_border()
        color = "006100" if grp_pnl > 0 else ("9C0006" if grp_pnl < 0 else "000000")
        c_pnl.font = Font(bold=True, color=color, name="微软雅黑", size=10)
        c_wr = ws.cell(row=ri_off, column=4, value=grp_wr)
        c_wr.number_format = "0.00%"
        c_wr.alignment = Alignment(horizontal="right", vertical="center")
        c_wr.fill = alt; c_wr.border = thin_border()
        ws.row_dimensions[ri_off].height = 18

    ws.column_dimensions["D"].width = 10


def write_sheet_by_symbol(wb: openpyxl.Workbook, positions: list[dict]):
    """按品种汇总Sheet"""
    ws = wb.create_sheet("按品种统计")

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="按品种统计（已平仓）")
    c.font      = Font(bold=True, size=13, color="FFFFFF", name="微软雅黑")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = make_fill(COLOR_HEADER)
    ws.row_dimensions[1].height = 28

    hdrs = ["品种", "总笔数", "盈利笔", "亏损笔", "净盈亏(点)", "胜率"]
    hdr_fill = make_fill(COLOR_SUBHEADER)
    for ci, h in enumerate(hdrs, 1):
        cell_style(ws, 2, ci, h, bold=True, font_color="FFFFFF",
                   fill=hdr_fill, align="center")

    closed = [p for p in positions if p["status"] != "open"]
    sym_map: dict[str, list] = {}
    for p in closed:
        sym_map.setdefault(p["symbol"], []).append(p)

    sym_stats = []
    for sym, plist in sym_map.items():
        wins = [p for p in plist if (p.get("pnl") or 0) > 0]
        pnl  = sum(p.get("pnl") or 0 for p in plist)
        sym_stats.append((sym, len(plist), len(wins),
                          len(plist) - len(wins), pnl,
                          len(wins) / len(plist) if plist else 0))

    sym_stats.sort(key=lambda x: x[4], reverse=True)

    for ri_off, (sym, total, w, l, pnl, wr) in enumerate(sym_stats, 3):
        alt = make_fill(COLOR_PROFIT if pnl > 0 else (COLOR_LOSS if pnl < 0 else COLOR_ALT_ROW))
        cell_style(ws, ri_off, 1, sym,   fill=alt, bold=True, align="center")
        cell_style(ws, ri_off, 2, total, fill=alt, align="center")
        cell_style(ws, ri_off, 3, w,     fill=alt, align="center")
        cell_style(ws, ri_off, 4, l,     fill=alt, align="center")
        c_pnl = ws.cell(row=ri_off, column=5, value=pnl)
        c_pnl.number_format = "+#,##0.00;-#,##0.00"
        c_pnl.alignment = Alignment(horizontal="right", vertical="center")
        c_pnl.fill = alt; c_pnl.border = thin_border()
        color = "006100" if pnl > 0 else ("9C0006" if pnl < 0 else "000000")
        c_pnl.font = Font(bold=True, color=color, name="微软雅黑", size=10)
        c_wr = ws.cell(row=ri_off, column=6, value=wr)
        c_wr.number_format = "0.00%"
        c_wr.alignment = Alignment(horizontal="right", vertical="center")
        c_wr.fill = alt; c_wr.border = thin_border()
        ws.row_dimensions[ri_off].height = 18

    widths = [10, 8, 8, 8, 14, 8]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_sheet_open(wb: openpyxl.Workbook, positions: list[dict]):
    """当前持仓Sheet"""
    ws = wb.create_sheet("当前持仓")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="当前持仓（open）")
    c.font      = Font(bold=True, size=13, color="FFFFFF", name="微软雅黑")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = make_fill(COLOR_HEADER)
    ws.row_dimensions[1].height = 28

    hdrs = [
        ("品种",8), ("方向",7), ("信号类型",9), ("入场时间",15),
        ("入场价",12), ("止损价",12), ("止盈价",12), ("风险距离",10),
        ("ATR",10), ("移动止损",9), ("交易ID",24), ("备注",10),
    ]
    hdr_fill = make_fill(COLOR_SUBHEADER)
    for ci, (h, w) in enumerate(hdrs, 1):
        cell_style(ws, 2, ci, h, bold=True, font_color="FFFFFF",
                   fill=hdr_fill, align="center")
        ws.column_dimensions[get_column_letter(ci)].width = w

    opens = [p for p in positions if p["status"] == "open"]
    opens_sorted = sorted(opens, key=lambda p: p.get("entryTime", ""))
    open_fill = make_fill(COLOR_OPEN)
    for ri_off, pos in enumerate(opens_sorted, 3):
        direc = pos.get("direction", "")
        fill = make_fill(COLOR_LONG if direc == "long" else COLOR_SHORT)

        def rc(col, val, num_fmt=None, align="center"):
            cell_style(ws, ri_off, col, val, fill=fill, align=align, num_fmt=num_fmt)

        rc(1,  pos.get("symbol", ""))
        rc(2,  DIR_MAP.get(direc, direc))
        rc(3,  SIG_MAP.get(pos.get("signalType",""), pos.get("signalType","")))
        rc(4,  pos.get("entryTime",""))
        rc(5,  pos.get("entryPrice"),  num_fmt="#,##0.00##")
        rc(6,  pos.get("stopLoss"),    num_fmt="#,##0.00##")
        rc(7,  pos.get("takeProfit"),  num_fmt="#,##0.00##")
        rc(8,  pos.get("riskDist"),    num_fmt="0.0000")
        rc(9,  pos.get("atr"),         num_fmt="0.0000")
        rc(10, "是" if pos.get("trailingActive") else "否")
        rc(11, pos.get("id",""), align="left")
        rc(12, "")
        ws.row_dimensions[ri_off].height = 18


def main():
    if not POSITIONS_FILE.exists():
        print(f"[ERROR] 找不到 positions.json: {POSITIONS_FILE}")
        return

    data     = json.loads(POSITIONS_FILE.read_text("utf-8"))
    positions = data.get("positions", [])
    print(f"[INFO] 共读取 {len(positions)} 笔交易记录")

    wb = openpyxl.Workbook()

    write_sheet_all(wb, positions)
    write_sheet_summary(wb, positions)
    write_sheet_by_symbol(wb, positions)
    write_sheet_open(wb, positions)

    wb.save(OUTPUT_FILE)
    print(f"[OK] 已导出至: {OUTPUT_FILE}")
    return OUTPUT_FILE


if __name__ == "__main__":
    main()
